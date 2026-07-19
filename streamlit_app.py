import streamlit as st
import pandas as pd
import pythoncom
import time
import subprocess
import os
import re
import unicodedata
from datetime import datetime, timedelta
from openpyxl import load_workbook
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import barcode
from barcode.writer import ImageWriter
import os

# ==============================================================================
# 1. FUNCIONES AUXILIARES DE PERSISTENCIA PARA STREAMLIT
# ==============================================================================

# --- CONFIGURACIÓN DE RUTAS PARA CATÁLOGOS ---
RUTA_CATALOGO_CLIENTES = r"C:\Users\BP_Juancarlos Palomi\Documents\Python_SAP\python-3.13.7-embed-amd64\catalogos\clientes_homologados.csv"
RUTA_CATALOGO_DESTINOS = r"C:\Users\BP_Juancarlos Palomi\Documents\Python_SAP\python-3.13.7-embed-amd64\catalogos\destinos_homologados.csv"
RUTA_EXCEL_CATALOGOS = r"C:\Users\BP_Juancarlos Palomi\Documents\Python_SAP\python-3.13.7-embed-amd64\catalogos\catalogos_homologacion.xlsx"

# Asegurar que la carpeta de catálogos exista
os.makedirs(os.path.dirname(RUTA_CATALOGO_CLIENTES), exist_ok=True)

def guardar_nuevo_registro_excel(pestaña, original, corto):
    """
    Agrega de forma segura un nuevo registro a una pestaña del catálogo de Excel,
    evitando duplicados y re-escribiendo de forma óptima el archivo.
    """
    try:
        # Cargar datos existentes
        df_completo = pd.read_excel(RUTA_EXCEL_CATALOGOS, sheet_name=None)
        df_pestana = df_completo[pestaña].copy()
        
        # Filtrar si ya existe el original (case insensitive) para evitar duplicados
        df_pestana = df_pestana[df_pestana["termino_busqueda"].astype(str).str.upper() != str(original).upper()]
        
        # Determinar columna destino según pestaña
        col_homologada = "cliente_homologado" if pestaña == "Clientes" else "destino_homologado"
        
        # Añadir el nuevo renglón
        nuevo_registro = pd.DataFrame([{"termino_busqueda": original, col_homologada: corto}])
        df_pestana = pd.concat([df_pestana, nuevo_registro], ignore_index=True)
        
        # Actualizar el diccionario de hojas
        df_completo[pestaña] = df_pestana
        
        # Guardar todas las hojas de vuelta en el Excel
        with pd.ExcelWriter(RUTA_EXCEL_CATALOGOS, engine="openpyxl") as writer:
            for sheet, df in df_completo.items():
                df.to_excel(writer, sheet_name=sheet, index=False)
                
        return True
    except Exception as e:
        st.error(f"Error al escribir en el catálogo Excel: {e}")
        return False

def cargar_catalogo(ruta):
    if os.path.exists(ruta):
        return pd.read_csv(ruta).to_dict(orient="records")
    return []

def guardar_catalogo(ruta, lista_datos):
    df = pd.DataFrame(lista_datos)
    df.to_csv(ruta, index=False, encoding="utf-8-sig")

def cargar_diccionarios_desde_excel(ruta_excel=RUTA_EXCEL_CATALOGOS):
    """
    Carga de forma dinámica las pestañas del archivo Excel y devuelve los diccionarios limpios.
    """
    try:
        df_clientes = pd.read_excel(ruta_excel, sheet_name="Clientes")
        df_destinos = pd.read_excel(ruta_excel, sheet_name="Destinos")
        df_cdfs = pd.read_excel(ruta_excel, sheet_name="CDFs")
        
        df_clientes = df_clientes.dropna(subset=["termino_busqueda"])
        df_destinos = df_destinos.dropna(subset=["termino_busqueda"])
        df_cdfs = df_cdfs.dropna(subset=["termino_busqueda"])
        
        dict_clientes_raw = {str(k).strip(): str(v).strip() for k, v in zip(df_clientes["termino_busqueda"], df_clientes["cliente_homologado"])}
        dict_destinos_raw = {str(k).strip(): str(v).strip() for k, v in zip(df_destinos["termino_busqueda"], df_destinos["destino_homologado"])}
        dict_cdfs_raw = {str(k).strip(): str(v).strip() for k, v in zip(df_cdfs["termino_busqueda"], df_cdfs["cdf_codigo"])}
        
        cli = dict(sorted(dict_clientes_raw.items(), key=lambda x: len(x[0]), reverse=True))
        dest = dict(sorted(dict_destinos_raw.items(), key=lambda x: len(x[0]), reverse=True))
        sec = dict(sorted(dict_cdfs_raw.items(), key=lambda x: len(x[0]), reverse=True))
        cdf = {k.lower(): v for k, v in sec.items() if "cdf" in v.lower()}
        
        return cli, dest, sec, cdf
    except Exception as e:
        st.warning(f"Aviso: No se pudieron cargar los catálogos desde Excel ({e}). Se usarán diccionarios vacíos.")
        return {}, {}, {}, {}

def actualizar_status_en_sheets(numero_viaje):
    """
    Busca el número de viaje (SHIPMENT) en la pestaña 'PLAN'
    y cambia su STATUS a 'FACTURADO' leyendo las credenciales directamente.
    """
    import os
    import streamlit as st
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        import json
        
        # Leemos las credenciales guardadas de forma segura en Streamlit Cloud
        info_credenciales = json.loads(st.secrets["llave_google"])
        
        SCOPES = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        
        # Conectamos usando el diccionario directamente en lugar de un archivo físico
        creds = Credentials.from_service_account_info(info_credenciales, scopes=SCOPES)
        cliente_sheets = gspread.authorize(creds)
        
        # ID de tu Google Sheet
        sheet_id = "1ES8JL8wTbh7ZUzNiCZpVQjzZEff565u2dz9oAR5maiE"
        sh = cliente_sheets.open_by_key(sheet_id)
        
        # Abrimos la pestaña "PLAN"
        worksheet = sh.worksheet("PLAN")
        
        # 2. Buscar en qué fila está el número de viaje (SHIPMENT) en la Columna B (Columna 2)
        celda = worksheet.find(str(numero_viaje), in_column=2)
        
        if celda:
            fila_encontrada = celda.row
            columna_status = 10  # Columna J (STATUS)
            
            # 3. Actualizar únicamente esa celda a "FACTURADO"
            worksheet.update_cell(fila_encontrada, columna_status, "FACTURADO")
            return True
        else:
            st.warning(f"⚠️ No se encontró el viaje {numero_viaje} en la pestaña PLAN.")
            return False
            
    except Exception as e:
        import traceback
        st.error("❌ Error detallado de Conexión:")
        st.code(traceback.format_exc(), language="text")
        return False

# ==============================================================================
# 🛠️ CONFIGURACIÓN DE PÁGINA GLOBAL
# ==============================================================================
st.set_page_config(
    page_title="Multiherramienta Logística SAP", 
    page_icon="🧰", 
    layout="wide"
)

# ==============================================================================
# 🧠 INICIALIZACIÓN DE ESTADOS DE SESIÓN (SESSION STATE)
# ==============================================================================
if 'entrega_num' not in st.session_state:
    st.session_state.entrega_num = ""
if 'tabla_validar' not in st.session_state:
    st.session_state.tabla_validar = None
if 'sap_variante_row' not in st.session_state:
    st.session_state.sap_variante_row = 301
if 'datos_order_flow' not in st.session_state:
    st.session_state.datos_order_flow = None
if 'df_flow_guardado' not in st.session_state:
    st.session_state.df_flow_guardado = None
if 'df_vt11_guardado' not in st.session_state:
    st.session_state.df_vt11_guardado = None

# --- CARGA SEGURA DE CATÁLOGOS EN SESSION STATE ---
if 'diccionario_clientes' not in st.session_state:
    cli, dest, sec, cdf = cargar_diccionarios_desde_excel()
    st.session_state.diccionario_clientes = cli
    st.session_state.diccionario_destinos = dest
    st.session_state.catalogo_cdfs_secundario = sec
    st.session_state.diccionario_cdf = cdf

if 'vt11_datos' not in st.session_state:
    st.session_state.vt11_datos = {
        "tipo_caja": "TR",
        "numero_caja": "",
        "peso": "26,000",
        "carta_porte": "",
        "placas": "",
        "chofer": "",
        "cortina": "CRT-   BN",
        "sello": "",
        "fecha_registro": "",
        "fecha_carga": "",
        "fecha_transporte": "",
        "fecha_planificacion": "",
        "hora_registro": "",
        "hora_carga": "",
        "hora_fin_carga": "",
        "hora_salida": "",
        "hora_transporte": "",
        "hora_planificacion": ""
    }

# Función auxiliar para recargar catálogos sin tocar variables globales
def recargar_catalogos_en_sesion():
    cli, dest, sec, cdf = cargar_diccionarios_desde_excel()
    st.session_state.diccionario_clientes = cli
    st.session_state.diccionario_destinos = dest
    st.session_state.catalogo_cdfs_secundario = sec
    st.session_state.diccionario_cdf = cdf

# ==============================================================================
# 2. FUNCIONES DE LIMPIEZA Y NORMALIZACIÓN
# ==============================================================================

def limpiar_texto(texto):
    if pd.isna(texto) or not isinstance(texto, str):
        return ""
    texto_limpio = texto.lower().strip()
    texto_limpio = "".join(
        c for c in unicodedata.normalize('NFD', texto_limpio)
        if unicodedata.category(c) != 'Mn'
    )
    texto_limpio = re.sub(r'\s+', ' ', texto_limpio)
    texto_limpio = texto_limpio.replace(",", "").replace(".", "").replace("´", "").replace("'", "")
    return texto_limpio.strip()

def normalizar_texto(texto):
    if not texto:
        return ""
    texto = str(texto).upper()
    texto = texto.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
    texto = re.sub(r'[-–—,_]', ' ', texto)
    return " ".join(texto.split())

# ==============================================================================
# 3. FUNCIONES DE HOMOLOGACIÓN Y BÚSQUEDA (CONECTADAS AL SESSION STATE)
# ==============================================================================

def buscar_cdf_por_coincidencia(destino_raw):
    destino_normalizado = normalizar_texto(destino_raw)
    if not destino_normalizado:
        return "CDF NO ENCONTRADO"
    
    catalogo_sec = st.session_state.get('catalogo_cdfs_secundario', {})
    if destino_normalizado in catalogo_sec:
        return catalogo_sec[destino_normalizado]
        
    palabras_destino = set(destino_normalizado.split())
    for llave_catalogo, cdf_resultado in catalogo_sec.items():
        llave_normalizada = normalizar_texto(llave_catalogo)
        palabras_catalogo = set(llave_normalizada.split())
        if palabras_catalogo.issubset(palabras_destino) or palabras_destino.issubset(palabras_catalogo):
            return cdf_resultado
            
    return f"REVISAR: {destino_raw}"

def obtener_cliente_corto(cliente_sap):
    if not cliente_sap or pd.isna(cliente_sap):
        return ""
        
    cliente_limpio = limpiar_texto(cliente_sap)
    dict_clientes = st.session_state.get('diccionario_clientes', {})
    
    if cliente_limpio in dict_clientes:
        return dict_clientes[cliente_limpio]
        
    for llave, cliente_corto in dict_clientes.items():
        if len(llave) > 3 and (llave in cliente_limpio or cliente_limpio in llave):
            return cliente_corto
            
    return str(cliente_sap).upper().strip()

def evaluar_cliente_relacion(nombre_sap):
    if not nombre_sap:
        return ""
    texto_sap_minusc = str(nombre_sap).strip().lower()
    dict_clientes = st.session_state.get('diccionario_clientes', {})
    for llave_catalogo, resultado_col2 in dict_clientes.items():
        if llave_catalogo in texto_sap_minusc:
            return resultado_col2
    return ""

def obtener_destino_corto(destino_sap, cliente_original):
    destino_limpio = str(destino_sap).strip()
    dict_destinos = st.session_state.get('diccionario_destinos', {})
    resultado = dict_destinos.get(destino_limpio)
    
    if resultado:
        return resultado
    else:
        return "N/A"

def obtener_cdf_soriana(destino_corto):
    destino_busqueda = limpiar_texto(destino_corto)
    dict_cdf = st.session_state.get('diccionario_cdf', {})
    for plaza, cdf_codigo in dict_cdf.items():
        if plaza in destino_busqueda:
            return cdf_codigo
            
    nombre_limpio = str(destino_corto).strip().upper()
    catalogo_sec = st.session_state.get('catalogo_cdfs_secundario', {})
    return catalogo_sec.get(nombre_limpio, "")


# ==============================================================================
# 2. FUNCIONES FÍSICAS DE ESCRITURA EN EXCEL (OPENPYXL)
# ==============================================================================

def generar_bitacora(ruta_plantilla, ruta_salida, df_vt11, df_viaje, firma_gdp):
    """
    Función optimizada que abre el molde BITACORA.xlsx, escribe los datos 
    extraídos de los dos DataFrames en las coordenadas solicitadas y lo guarda.
    Ajustado para impresión en una sola hoja con márgenes de 2.5cm arriba/abajo y 1.9cm lados.
    """
    try:
        # 1. Cargar la plantilla Excel existente
        wb = openpyxl.load_workbook(ruta_plantilla)
        ws = wb.active  # Selecciona la hoja activa del molde
        
        # 2. Asegurar que tenemos datos para leer en los DataFrames
        fila_vt11 = df_vt11.iloc[0] if df_vt11 is not None and not df_vt11.empty else {}
        fila_viaje = df_viaje.iloc[0] if df_viaje is not None and not df_viaje.empty else {}
        
        # === SECCIÓN 1: GENERALES ===
        ws['E7'] = fila_vt11.get('Name of service provider', '')
        ws['D8'] = fila_vt11.get('Route', '')
        ws['I8'] = fila_vt11.get('Descrip.of Shipment', '')
        ws['D9'] = fila_vt11.get('Shipment', '')
        ws['I9'] = fila_vt11.get('External ID 2', '')
        ws['D10'] = fila_vt11.get('Route name', '')
        ws['I10'] = fila_vt11.get('Container ID', '')
        
        # === SECCIÓN 2: REGISTRO DE TIEMPOS ===
        # 1. Cita de Carga (Planificado Check-in)
        ws['E14'] = fila_vt11.get('PlDat.ChIn', '')
        ws['I14'] = fila_vt11.get('PlTmChIn', '')
        
        # 2. Llegada a Cita de Carga (Real Check-in)
        ws['E15'] = fila_vt11.get('ActDteChIn', '')
        ws['I15'] = fila_vt11.get('CurTmChI', '')
        
        # 3. Entrada a Cargar (Inicio de Carga)
        ws['E16'] = fila_vt11.get('CurrLoadSt', '')
        ws['I16'] = fila_vt11.get('AcLdST', '')
        
        # 4. Entrega de Documentos (Fin de Carga)
        ws['E17'] = fila_vt11.get('ActLoadEnd', '')
        ws['I17'] = fila_vt11.get('AcLdET', '')
        
        # 5. Salida del CEDIS (Salida de Transporte)
        ws['E18'] = fila_vt11.get('CurrShipSt', '')
        ws['I18'] = fila_vt11.get('AcTrST', '')
        
        # === SECCIÓN 3: CITA DE DESCARGA, OBSERVACIONES Y CLIENTE ===
        ws['E20'] = fila_viaje.get('Fecha App', '')
        ws['J20'] = fila_viaje.get('Confirma SAP', '')
        ws['J21'] = fila_viaje.get('Hora App', '')
        
        ws['E25'] = fila_viaje.get('Comentarios', '')
        ws['K25'] = fila_vt11.get('SpPl', '')
        
        # --- DETECCIÓN INTELIGENTE DE CLIENTE CORTO + DESTINO (E27) ---
        cliente_sap = str(fila_viaje.get('Cliente', '')).upper()
        destino_raw = fila_vt11.get('Route name', '')

        # 1. Traducir el cliente con tu diccionario
        cliente_corto = obtener_cliente_corto(cliente_sap)

        # Si devolvió el nombre largo (sin traducir), aplicamos el filtro de limpieza
        if cliente_corto == cliente_sap or "CENTRO DE DISTRIBU" in cliente_corto:
            for remover in ["CENTRO DE DISTRIBUCIÓN", "CENTRO DE DISTRIBUCION", "CENTRO DISTRIBUCION", "CEDIS"]:
                cliente_corto = cliente_corto.replace(remover, "").strip()
            
            if cliente_corto.startswith("DE "):
                cliente_corto = cliente_corto[3:].strip()
            
            # Unificaciones comunes
            if "SAMS" in cliente_corto:
                cliente_corto = "SAMS"
            elif "WALMART" in cliente_corto or "WAL-MART" in cliente_corto:
                cliente_corto = "WALMART"
            elif "SORIANA" in cliente_corto:
                cliente_corto = "SORIANA"
            elif "CHEDRAUI" in cliente_corto:
                cliente_corto = "CHEDRAUI"

        # 2. Traducir destino usando tu diccionario
        destino_corto_traducido = obtener_destino_corto(destino_raw, cliente_sap)

        # Limpieza final del destino traducido
        destino_limpio = str(destino_corto_traducido).strip().rstrip(",")

        # 3. Escribimos el resultado en E27
        if cliente_corto and destino_limpio:
            ws['E27'] = f"{cliente_corto} - {destino_limpio}"
        elif cliente_corto:
            ws['E27'] = cliente_corto
        else:
            ws['E27'] = destino_limpio
        
        # === SECCIÓN 4: FIRMAS ===
        # --- NORMALIZACIÓN DE LA FIRMA GDP (G37) ---
        firma_usuario = str(firma_gdp).strip().upper()
        
        # Si detectamos "MARTIN" o variantes, lo cambiamos a "M.NARVAEZ R."
        if firma_usuario in ["MARTIN", "MARTIN NARVAEZ", "MARTIN NARVÁEZ", "M. NARVAEZ"]:
            firma_final = "M.NARVAEZ R."
        elif not firma_usuario:  # Si viene vacío por error, le asignamos el suyo por defecto
            firma_final = "M.NARVAEZ R."
        else:
            firma_final = str(firma_gdp).strip() # Si firma otro operador, respeta su nombre

        ws['G37'] = firma_final  # Escribe la firma limpia en G37
        ws['B49'] = fila_vt11.get('Descrip.of Shipment', '') # Firma del operador en B49
        
        # === SECCIÓN DE MAQUETADO DE IMPRESIÓN (MÁRGENES Y AJUSTE) ===
        # Configurar ajuste dinámico a una sola hoja
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        
        # Asignar márgenes correspondientes (2.5 cm arriba/abajo y 1.9 cm lados)
        ws.page_margins.top = 0.984      
        ws.page_margins.bottom = 0.984   
        ws.page_margins.left = 0.748     
        ws.page_margins.right = 0.748    

        # 3. Asegurar que la carpeta de destino exista
        os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
        
        # 4. Guardar archivo final
        wb.save(ruta_salida)
        wb.close()
        return True

    except Exception as e:
        st.error(f"❌ Error al procesar la plantilla de Bitácora: {str(e)}")
        return False


def generar_pase_vehicular(fila_vt11, datos_viaje, lista_pos, lista_facturas, numero_viaje):
    """
    Carga la plantilla 'PV' de la carpeta de plantillas, la rellena con 
    la información del viaje actual y la guarda como un archivo individual.
    Ajustado para impresión en una sola hoja con márgenes de 2.5cm arriba/abajo y 1.9cm lados.
    """
    # 1. Definir rutas de archivos
    ruta_plantilla = r"C:\Users\BP_Juancarlos Palomi\Documents\Python_SAP\python-3.13.7-embed-amd64\plantillas\PV.xlsx"
    
    # Carpeta donde se guardarán los Pases Vehiculares generados (la crea si no existe)
    carpeta_salida = r"C:\Users\BP_Juancarlos Palomi\Documents\Python_SAP\python-3.13.7-embed-amd64\pases_generados"
    if not os.path.exists(carpeta_salida):
        os.makedirs(carpeta_salida)
        
    ruta_salida = os.path.join(carpeta_salida, f"PV_Viaje_{numero_viaje}.xlsx")
    
    # 2. Cargar el libro de Excel de la plantilla
    try:
        wb_pv = openpyxl.load_workbook(ruta_plantilla)
        ws_pv = wb_pv.active 
    except Exception as e:
        print(f"❌ Error al cargar la plantilla PV: {e}")
        return

    # 3. Llenar los campos básicos del operador, fecha y cita
    ws_pv['C30'] = fila_vt11.get('Descrip.of Shipment', '')
    ws_pv['B32'] = datos_viaje.get('fecha_app', '')
    ws_pv['E32'] = datos_viaje.get('hora_app', '')
    ws_pv['G32'] = datos_viaje.get('confirma_sap', '')
    
    # 4. Llenar placas de transporte
    ws_pv['C34'] = datos_viaje.get('', '') 
    ws_pv['F34'] = fila_vt11.get('External ID 2', '') # Placa Remolque
    
    # 5. Llenado dinámico de Orden de Compra (PO) en D38:D46 y Factura en G38:G46
    max_filas = 9 # De la fila 38 a la 46 hay exactamente 9 espacios
    
    # Escribir POs únicas
    for i, po in enumerate(lista_pos[:max_filas]):
        celda_po = f"D{38 + i}"
        ws_pv[celda_po] = po
        
    # Escribir Facturas únicas
    for i, factura in enumerate(lista_facturas[:max_filas]):
        celda_factura = f"G{38 + i}"
        ws_pv[celda_factura] = factura

    # === SECCIÓN DE MAQUETADO DE IMPRESIÓN (MÁRGENES Y AJUSTE) ===
    # Configurar ajuste dinámico a una sola hoja
    ws_pv.sheet_properties.pageSetUpPr.fitToPage = True
    ws_pv.page_setup.fitToWidth = 1
    ws_pv.page_setup.fitToHeight = 1
    
    # Asignar márgenes correspondientes al Pase (2.5 cm arriba/abajo y 1.9 cm lados)
    ws_pv.page_margins.top = 0.984      
    ws_pv.page_margins.bottom = 0.984   
    ws_pv.page_margins.left = 0.748     
    ws_pv.page_margins.right = 0.748    

    # 6. Guardar el archivo final
    try:
        wb_pv.save(ruta_salida)
        print(f"✅ Pase Vehicular generado con éxito para el viaje {numero_viaje}: {ruta_salida}")
    except Exception as e:
        print(f"❌ Error al guardar el Pase Vehicular del viaje {numero_viaje}: {e}")
    finally:
        wb_pv.close()


def generar_es(plantilla_path, destino_path, datos):
    """
    Escribe los datos correspondientes en la plantilla de ES.xlsx (Soriana)
    Ajustado para impresión en una sola hoja con márgenes de 1.905cm arriba/abajo y 1.778cm lados.
    """
    wb = load_workbook(plantilla_path)
    ws = wb.active
    
    ws["C12"] = datos["cita"]  # Confirmación SAP
    ws["E5"] = datos["fecha_app"]
    ws["E6"] = datos["hora_app"]
    ws["C16"] = datos["operador"]
    ws["C19"] = datos["cdf"]  # CDF de Soriana (Ej: "5548 CDF TULTITLAN")
    
    # === SECCIÓN DE MAQUETADO DE IMPRESIÓN (MÁRGENES Y AJUSTE) ===
    # Configurar ajuste dinámico a una sola hoja
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    
    # Asignar márgenes correspondientes al Pase E/S (1.905 cm arriba/abajo y 1.778 cm lados)
    ws.page_margins.top = 0.75         
    ws.page_margins.bottom = 0.75      
    ws.page_margins.left = 0.70        
    ws.page_margins.right = 0.70       
    
    wb.save(destino_path)
    wb.close()

# =====================================================================
# 🛠️ GLOBAL NAVIGATION (SIDEBAR)
# =====================================================================
with st.sidebar:
    st.title("🧰 Menú Central")
    st.subheader("Selecciona un Panel")
    
    # Enrutador de pestañas de paneles
    opcion = st.selectbox(
        "Herramienta activa:",
        ["🚛 SAP Logistics Master", "⚡ Control Tránsitos MX21"]
    )
    st.divider()


# =====================================================================
# 🚛 HERRAMIENTA 1: SAP LOGISTICS MASTER (Código Completo)
# =====================================================================
def formatear_fecha_h1(texto_fecha):
    texto = ''.join(filter(str.isdigit, str(texto_fecha)))
    if len(texto) == 6:
        return f"{texto[0:2]}.{texto[2:4]}.20{texto[4:6]}"
    elif len(texto) == 8:
        return f"{texto[0:2]}.{texto[2:4]}.{texto[4:8]}"
    return texto_fecha

def formatear_hora_h1(texto_hora):
    texto = ''.join(filter(str.isdigit, str(texto_hora)))
    if len(texto) == 3:
        texto = "0" + texto
    if len(texto) == 4:
        return f"{texto[0:2]}:{texto[2:4]}"
    return texto_hora

def conectar_sap_h1():
    pythoncom.CoInitialize()
    try:
        sap_gui = win32com.client.GetObject("SAPGUI")
        app = sap_gui.GetScriptingEngine
        connection = app.Children(0)
        session = connection.Children(0)
        return session
    except Exception as e:
        st.error(f"❌ No se pudo conectar a SAP. Asegúrate de que SAP GUI esté abierto. Error: {e}")
        return None

def verificar_y_bajar_status_sap(shipment_id):
    if not shipment_id:
        return
    session = conectar_sap_h1()
    if session:
        try:
            session.findById("wnd[0]/tbar[0]/okcd").text = "/nVT11"
            session.findById("wnd[0]").sendVKey(0)
            session.findById("wnd[0]/usr/ctxtK_TKNUM-LOW").text = shipment_id
            session.findById("wnd[0]").sendVKey(8)
            session.findById("wnd[0]/usr/lbl[8,4]").setFocus()
            session.findById("wnd[0]").sendVKey(2)
            
            session.findById("wnd[0]/usr/tabsHEADER_TABSTRIP2/tabpTABS_OV_DE").select()
            tp = "wnd[0]/usr/tabsHEADER_TABSTRIP2/tabpTABS_OV_DE/ssubG_HEADER_SUBSCREEN2:SAPMV56A:1025/"
            
            bajo_status = False
            try:
                session.findById(tp + "btn*RV56A-ICON_STLBG").press()
                bajo_status = True
            except: pass
                
            try:
                session.findById(tp + "btn*RV56A-ICON_STREG").press()
                bajo_status = True
            except: pass
            
            if bajo_status:
                session.findById("wnd[0]/tbar[0]/btn[11]").press()
                time.sleep(1)
                while session.Children.Count > 1:
                    session.findById("wnd[1]/tbar[0]/btn[0]").press()
                    time.sleep(0.5)
                st.toast(f"🔄 Shipment {shipment_id} detectado en Status alto. ¡Se bajó a Status 1 automáticamente! 🔓", icon="⚠️")
            else:
                st.toast(f"✅ Shipment {shipment_id} validado correctamente. Ya está libre en Status 1.", icon="👍")
                session.findById("wnd[0]/tbar[0]/okcd").text = "/n"
                session.findById("wnd[0]").sendVKey(0)
        except Exception as e:
            st.toast(f"ℹ️ Modo manual / Sin conexión SAP activa.", icon="💻")
        finally:
            pythoncom.CoUninitialize()

@st.dialog("⚙️ Cambiar Fila Objetivo")
def ventana_cambiar_numero_fila():
    st.markdown("### Ajustar Fila de Variante")
    st.write("Modifica manualmente el número de fila objetivo si detectas un desplazamiento en la auditoría:")
    nueva_fila = st.number_input(
        "Fila de la Variante en SAP (v_grid.currentCellRow)", 
        min_value=0, 
        value=int(st.session_state.sap_variante_row),
        step=1
    )
    st.divider()
    if st.button("💾 Aplicar Cambios", type="primary", use_container_width=True):
        st.session_state.sap_variante_row = nueva_fila
        st.success(f"¡Fila actualizada a: {nueva_fila}!")
        time.sleep(0.8)
        st.rerun()

def render_logistica_master():
    PERFILES_USUARIOS = {
        "MARTIN": {"tripulacion": "7", "add04": "PICK010", "facturista": "MARTIN", "texto4": "BN MARTIN", "gdp": "M.NARVAEZ R."},
        "OPERADOR_2": {"tripulacion": "5", "add04": "PICK020", "facturista": "JUAN LUIS", "texto4": "BN JUAN", "gdp": "JUAN LUIS"}
    }
    WALMART_MATERIAL = "HMXTAR04"
    WALMART_CENTRO = "MX21"
    WALMART_ALMACEN = "FG"

    with st.sidebar:
        st.subheader("⚙️ Configuración Módulo")
        usuario_sel = st.selectbox("👤 Perfil de Facturación (Segundo Plano)", list(PERFILES_USUARIOS.keys()))
        perfil = PERFILES_USUARIOS[usuario_sel]
        
        if st.button("⚙️ Editar Número de Fila", use_container_width=True):
            ventana_cambiar_numero_fila()
        
        st.divider()
        st.header("🔑 Shipment Activo")
        entrega = st.text_input("Número de Entrega / Transporte", value=st.session_state.entrega_num, key="input_shipment")
        bajar_status_permitido = st.checkbox("¿Bajar Estatus automáticamente en SAP?", value=False)
        
        if entrega != st.session_state.entrega_num:
            st.session_state.entrega_num = entrega
            if bajar_status_permitido:
                verificar_y_bajar_status_sap(entrega)
            else:
                st.toast(f"ℹ️ Shipment {entrega} cargado sin modificar estatus (Modo Seguro).", icon="🔒")
            
        st.divider()
        st.subheader("📦 Paso 1: Walmart")
        f1_cant = st.text_input("Cantidad", value="")

        if st.button("🚀 EJECUTAR Walmart", use_container_width=True, type="primary"):
            # --- 🧹 RUTINA DE AUTO-LIMPIEZA DE BASURA TEMPORAL ---
            try:
                import glob
                carpetas_a_limpiar = [
                    r"C:\Users\BP_Juancarlos Palomi\Documents\Python_SAP\python-3.13.7-embed-amd64\pases_generados",
                    r"C:\Users\BP_Juancarlos Palomi\Documents\Python_SAP\python-3.13.7-embed-amd64\documentos_generados"
                ]
                
                archivos_eliminados = 0
                for carpeta in carpetas_a_limpiar:
                    if os.path.exists(carpeta):
                        # Buscamos y borramos únicamente archivos de Excel y PDF viejos
                        for extension in ["*.xlsx", "*.pdf"]:
                            for archivo in glob.glob(os.path.join(carpeta, extension)):
                                try:
                                    os.remove(archivo)
                                    archivos_eliminados += 1
                                except Exception:
                                    # Si el archivo está abierto en Excel o PDF Reader, lo saltamos silenciosamente
                                    pass
                
                if archivos_eliminados > 0:
                    st.sidebar.info(f"🧹 Se eliminaron {archivos_eliminados} archivos temporales del viaje anterior.")
            except Exception as e_clean:
                st.sidebar.warning(f"⚠️ No se pudo completar la limpieza de temporales: {e_clean}")
            
            # --- ⚙️ INICIO DEL PROCESO SAP ---
            try:
                pythoncom.CoInitialize()
                session = win32com.client.GetObject("SAPGUI").GetScriptingEngine.Children(0).Children(0)
                session.findById("wnd[0]").maximize()
                session.findById("wnd[0]/tbar[0]/okcd").text = "/nY_LAD_65000730"
                session.findById("wnd[0]").sendVKey(0)
                time.sleep(1)
                session.findById("wnd[0]/usr/ctxtPLN_PNT-LOW").text = WALMART_CENTRO
                session.findById("wnd[0]/usr/ctxtSHP_STS-LOW").text = "0"
                session.findById("wnd[0]/usr/ctxtSHP_STS-HIGH").text = "6"
                session.findById("wnd[0]/usr/ctxtSHP_NUM-LOW").text = st.session_state.entrega_num
                session.findById("wnd[0]").sendVKey(8)
                time.sleep(1)
                session.findById("wnd[0]/usr/lbl[26,8]").setFocus()
                session.findById("wnd[0]").sendVKey(2)
                time.sleep(1)
                session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\02").select()
                session.findById("wnd[0]/tbar[1]/btn[25]").press()
                time.sleep(3)
                
                tbl_path = "wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\02/ssubSUBSCREEN_BODY:SAPMV50A:1104/tblSAPMV50ATC_LIPS_PICK"
                success = False
                for row in range(0, 11):
                    try:
                        campo_mat = session.findById(f"{tbl_path}/ctxtLIPS-MATNR[1,{row}]")
                        if campo_mat.changeable:
                            campo_mat.text = WALMART_MATERIAL
                            session.findById(f"{tbl_path}/ctxtLIPS-WERKS[2,{row}]").text = WALMART_CENTRO
                            session.findById(f"{tbl_path}/ctxtLIPS-LGORT[3,{row}]").text = WALMART_ALMACEN
                            session.findById(f"{tbl_path}/txtLIPSD-G_LFIMG[5,{row}]").text = f1_cant
                            success = True
                            break
                    except: continue
                if success:
                    session.findById("wnd[0]").sendVKey(0)
                    time.sleep(1.5)
                    session.findById("wnd[0]").sendVKey(11)
                    time.sleep(1.5)
                    st.sidebar.success(f"✅ Guardado en fila {row}")
                else:
                    st.sidebar.error("❌ No se encontró fila modificable.")
                session.findById("wnd[0]").sendVKey(3)
                session.findById("wnd[0]").sendVKey(3)
            except Exception as e: st.sidebar.error(f"Error: {e}")
            finally: pythoncom.CoUninitialize()

        st.divider()
        st.subheader("📄 Paso 2: Documentos")
        if st.button("🖨️ Walmart Spool", use_container_width=True):
            try:
                pythoncom.CoInitialize()
                session = win32com.client.GetObject("SAPGUI").GetScriptingEngine.Children(0).Children(0)
                session.findById("wnd[0]/tbar[0]/okcd").text = "/nY_LAD_65000730"
                session.findById("wnd[0]").sendVKey(0)
                time.sleep(1)
                session.findById("wnd[0]/usr/ctxtPLN_PNT-LOW").text = WALMART_CENTRO
                session.findById("wnd[0]/usr/ctxtSHP_STS-LOW").text = "0"
                session.findById("wnd[0]/usr/ctxtSHP_STS-HIGH").text = "6"
                session.findById("wnd[0]/usr/ctxtSHP_NUM-LOW").text = st.session_state.entrega_num
                session.findById("wnd[0]").sendVKey(8)
                time.sleep(1)
                session.findById("wnd[0]/tbar[0]/btn[86]").press()
                time.sleep(1.5)
                try:
                    session.findById("wnd[1]/usr/subSUBSCREEN:SAPLSPRI:0600/cmbPRIPAR_DYN-PRIMM").key = ""
                    session.findById("wnd[1]/tbar[0]/btn[13]").press()
                    st.success("✅ Enviado a Spool.")
                except: st.error("❌ Error ventana impresión.")
                time.sleep(1)
                session.findById("wnd[0]/tbar[0]/btn[3]").press()
                session.findById("wnd[0]/tbar[0]/okcd").text = "/n"
                session.findById("wnd[0]").sendVKey(0)
            except Exception as e: st.error(f"Error Spool: {e}")
            finally: pythoncom.CoUninitialize()

    st.title("🚛 Centro de Mando MX21 SAP - V8.1")

# =====================================================================
# 🛠️ VL06O
# =====================================================================
    
    tab_vl06o, tab_vt11, tab_order_flow, tab_formatos_excel, tab_spool = st.tabs([
        "VL06O", 
        "VT11 Smart & Documentos", 
        "ORDER FLOW", 
        "FORMATOS EXCEL", 
        "Spool"
    ])

    with tab_vl06o:
        st.header("VL06O")
        tknum_vl = st.session_state.get('entrega_num', "")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🛰️ 1. Extraer Datos de SAP", use_container_width=True, key="vl06_btn1"):
                session = conectar_sap_h1()
                if session:
                    try:
                        session.findById("wnd[0]/tbar[0]/okcd").text = "/nVL06O"
                        session.findById("wnd[0]").sendVKey(0)
                        session.findById("wnd[0]/usr/btnBUTTON4").press()
                        session.findById("wnd[0]/usr/ctxtIT_WADAT-LOW").text = ""
                        session.findById("wnd[0]/usr/ctxtIT_WADAT-HIGH").text = ""
                        session.findById("wnd[0]/usr/ctxtIT_TKNUM-LOW").text = tknum_vl
                        session.findById("wnd[0]").sendVKey(8)
                        time.sleep(1.5)
                        session.findById("wnd[0]/tbar[1]/btn[18]").press()
                        session.findById("wnd[0]/tbar[1]/btn[45]").press()
                        try:
                            session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[5,0]").select()
                            session.findById("wnd[1]/tbar[0]/btn[0]").press()
                        except: pass
                        time.sleep(1)
                        df = pd.read_clipboard(sep='\t')
                        df = df.dropna(how='all', axis=1)
                        st.session_state.tabla_validar = df
                        session.findById("wnd[0]/tbar[1]/btn[5]").press()
                        st.success("✅ Datos capturados de VL06O.")
                    except Exception as e: st.error(f"❌ Error en SAP: {e}")

        if st.session_state.tabla_validar is not None:
            df = st.session_state.tabla_validar
            def highlight_sap(row):
                is_total = any('*' in str(v) for v in row.values)
                return ['background-color: #ffff00; color: black; font-weight: bold' if is_total else '' for _ in row]
            st.dataframe(df.style.apply(highlight_sap, axis=1), use_container_width=True)
            with col2:
                if st.button("✅ 2. Confirmar y dar F8 en SAP", type="primary", use_container_width=True, key="vl06_btn2"):
                    session = conectar_sap_h1()
                    if session:
                        try:
                            session.findById("wnd[0]").sendVKey(8)
                            time.sleep(1)
                            try: session.findById("wnd[1]").sendVKey(0)
                            except: pass
                            st.balloons()
                            st.success("🚀 Proceso completado exitosamente.")
                            st.session_state.tabla_validar = None
                        except Exception as e: st.error(f"Error al confirmar: {e}") 

# =====================================================================
# 🛠️VT11
# =====================================================================

    with tab_vt11:
        st.subheader("Procesamiento Individual VT11")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown("### 🚛 Identificación Esencial")
            sub_c1, sub_c2 = st.columns([1, 2])
            with sub_c1: prefijo = st.selectbox("Tipo", ["TR", "CA", "TH"], index=0)
            with sub_c2: v_caja_num = st.text_input("Número Caja", st.session_state.vt11_datos['numero_caja'])
            
            v_peso = st.text_input("Peso (TWGT)", st.session_state.vt11_datos['peso'])
            v_exti1 = st.text_input("Carta porte", st.session_state.vt11_datos['carta_porte'])
            v_exti2 = st.text_input("Placas ", st.session_state.vt11_datos['placas'])
            v_chofer = st.text_input("Chofer", st.session_state.vt11_datos['chofer'])
            v_remolque = st.text_input("Cortina", st.session_state.vt11_datos['cortina'])
            v_text1 = st.text_input("Sello", st.session_state.vt11_datos['sello'])
            
            v_add02 = perfil["tripulacion"]
            v_add04 = perfil["add04"]
            v_text2 = perfil["facturista"]
            v_text3 = "5"
            v_text4 = perfil["texto4"]

        with col2:
            st.markdown("### 📅 Fechas (Solo Números)")
            f1_raw = st.text_input("F. Registro", st.session_state.vt11_datos['fecha_registro'])
            f2_raw = st.text_input("F. Inicio Carga", st.session_state.vt11_datos['fecha_carga'])
            f3_raw = st.text_input("F. Inicio Transp", st.session_state.vt11_datos['fecha_transporte'])
            f4_raw = st.text_input("F. Planificación", st.session_state.vt11_datos['fecha_planificacion'])
            
            f1 = formatear_fecha_h1(f1_raw)
            f2 = formatear_fecha_h1(f2_raw)
            f3 = formatear_fecha_h1(f3_raw)
            f4 = formatear_fecha_h1(f4_raw)

        with col3:
            st.markdown("### ⏰ Horas (Solo Números)")
            h1_raw = st.text_input("H. Registro", st.session_state.vt11_datos['hora_registro'])
            h2_raw = st.text_input("H. Inicio Carga", st.session_state.vt11_datos['hora_carga'])
            h3_raw = st.text_input("H. Fin Carga", st.session_state.vt11_datos['hora_fin_carga'])
            h_sal_raw = st.text_input("H. Salida", st.session_state.vt11_datos['hora_salida'])
            h_it_raw = st.text_input("H. Inicio Transp", st.session_state.vt11_datos['hora_transporte'])
            h4_raw = st.text_input("H. Planificación", st.session_state.vt11_datos['hora_planificacion'])
            
            h1 = formatear_hora_h1(h1_raw)
            h2 = formatear_hora_h1(h2_raw)
            h3 = formatear_hora_h1(h3_raw)
            h_sal = formatear_hora_h1(h_sal_raw)
            h_it = formatear_hora_h1(h_it_raw)
            h4 = formatear_hora_h1(h4_raw)

        st.divider()
        btn_col1, btn_col2 = st.columns(2)
        
        with btn_col1:
            if st.button("📥 PRESIONAR PARA LLENAR SAP (VT11)", use_container_width=True, type="secondary", key="btn_llenar_sap"):
                caja_completa = f"{prefijo}-{v_caja_num}"
                try:
                    pythoncom.CoInitialize()
                    session = win32com.client.GetObject("SAPGUI").GetScriptingEngine.Children(0).Children(0)
                    
                    session.findById("wnd[0]/tbar[0]/okcd").text = "/nVT11"
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById("wnd[0]/usr/ctxtK_TKNUM-LOW").text = st.session_state.entrega_num
                    session.findById("wnd[0]").sendVKey(8)
                    session.findById("wnd[0]/usr/lbl[8,4]").setFocus()
                    session.findById("wnd[0]").sendVKey(2)
                    
                    session.findById("wnd[0]/usr/tabsHEADER_TABSTRIP1/tabpTABS_OV_PR").select()
                    session.findById("wnd[0]/usr/tabsHEADER_TABSTRIP1/tabpTABS_OV_PR/ssubG_HEADER_SUBSCREEN1:SAPMV56A:1021/txtVTTK-SIGNI").text = caja_completa
                    session.findById("wnd[0]/usr/tabsHEADER_TABSTRIP1/tabpTABS_OV_PR/ssubG_HEADER_SUBSCREEN1:SAPMV56A:1021/txtVTTK-ALLOWED_TWGT").text = v_peso
                    session.findById("wnd[0]/usr/tabsHEADER_TABSTRIP1/tabpTABS_OV_PR/ssubG_HEADER_SUBSCREEN1:SAPMV56A:1021/ctxtVTTK-EXTI1").text = v_exti1
                    
                    session.findById("wnd[0]/usr/tabsHEADER_TABSTRIP1/tabpTABS_OV_ID").select()
                    if session.Children.Count > 1: session.findById("wnd[1]/tbar[0]/btn[0]").press()
                    session.findById("wnd[0]/usr/tabsHEADER_TABSTRIP1/tabpTABS_OV_ID/ssubG_HEADER_SUBSCREEN1:SAPMV56A:1022/txtVTTK-EXTI2").text = v_exti2
                    session.findById("wnd[0]/usr/tabsHEADER_TABSTRIP1/tabpTABS_OV_ID/ssubG_HEADER_SUBSCREEN1:SAPMV56A:1022/txtVTTK-TPBEZ").text = v_chofer
                    session.findById("wnd[0]/usr/tabsHEADER_TABSTRIP1/tabpTABS_OV_ID/ssubG_HEADER_SUBSCREEN1:SAPMV56A:1022/txtVTTK-TNDR_TRKID").text = v_remolque
                    
                    session.findById("wnd[0]/usr/tabsHEADER_TABSTRIP2/tabpTABS_OV_AI").select()
                    ai = "wnd[0]/usr/tabsHEADER_TABSTRIP2/tabpTABS_OV_AI/ssubG_HEADER_SUBSCREEN2:SAPMV56A:1030/"
                    session.findById(ai + "ctxtVTTK-ADD02").text = v_add02
                    session.findById(ai + "ctxtVTTK-ADD04").text = v_add04
                    session.findById(ai + "txtVTTK-TEXT1").text = v_text1
                    session.findById(ai + "txtVTTK-TEXT2").text = v_text2
                    session.findById(ai + "txtVTTK-TEXT3").text = v_text3
                    session.findById(ai + "txtVTTK-TEXT4").text = v_text4
                    
                    session.findById("wnd[0]/usr/tabsHEADER_TABSTRIP2/tabpTABS_OV_DE").select()
                    tp = "wnd[0]/usr/tabsHEADER_TABSTRIP2/tabpTABS_OV_DE/ssubG_HEADER_SUBSCREEN2:SAPMV56A:1025/"
                    session.findById(tp + "btn*RV56A-ICON_STREG").press()
                    if session.Children.Count > 1: session.findById("wnd[1]/tbar[0]/btn[8]").press()
                    session.findById(tp + "btn*RV56A-ICON_STLBG").press()
                    session.findById(tp + "btn*RV56A-ICON_STLAD").press()
                    session.findById(tp + "btn*RV56A-ICON_STABF").press()
                    session.findById(tp + "btn*RV56A-ICON_STTBG").press()
                    if session.Children.Count > 1: session.findById("wnd[1]/tbar[0]/btn[8]").press()
                    if session.Children.Count > 1: session.findById("wnd[2]/tbar[0]/btn[0]").press()
                    
                    session.findById(tp + "ctxtVTTK-UAREG").text = h1
                    session.findById(tp + "ctxtVTTK-UALBG").text = h2
                    session.findById(tp + "ctxtVTTK-UALEN").text = h3
                    session.findById(tp + "ctxtVTTK-UZABF").text = h_sal
                    session.findById(tp + "ctxtVTTK-UATBG").text = h_it
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById(tp + "ctxtVTTK-DALBG").text = f2
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById(tp + "ctxtVTTK-DAREG").text = f1
                    session.findById(tp + "ctxtVTTK-DATBG").text = f3
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById(tp + "ctxtVTTK-DPTBG").text = f4
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById(tp + "ctxtVTTK-UPTBG").text = h4
                    session.findById("wnd[0]").sendVKey(0)
                    
                    session.findById("wnd[0]/tbar[0]/btn[11]").press()
                    time.sleep(1.2)
                    while session.Children.Count > 1:
                        session.findById("wnd[1]/tbar[0]/btn[0]").press()
                        time.sleep(0.5)
                    session.findById("wnd[0]/tbar[0]/okcd").text = "/n"
                    session.findById("wnd[0]").sendVKey(0)
                    st.success("✅ Datos inyectados y guardados en SAP exitosamente.")
                except Exception as e: st.error(f"❌ Error al interactuar con SAP: {e}")
                finally: pythoncom.CoUninitialize()

        with btn_col2:
            if st.button("📄 GENERAR TXT DE VT11", use_container_width=True, type="primary", key="btn_vt11_txt"):
                try:
                    pythoncom.CoInitialize()
                    session = win32com.client.GetObject("SAPGUI").GetScriptingEngine.Children(0).Children(0)
                    session.findById("wnd[0]/tbar[0]/okcd").text = "/nVT11"
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById("wnd[0]/usr/ctxtK_TKNUM-LOW").text = st.session_state.entrega_num
                    session.findById("wnd[0]").sendVKey(8)
                    session.findById("wnd[0]/mbar/menu[0]/menu[6]/menu[2]").select()
                    session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").select()
                    session.findById("wnd[1]/tbar[0]/btn[0]").press()
                    nombre_txt = st.session_state.entrega_num[-4:] if len(st.session_state.entrega_num) >= 4 else "repo"
                    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = f"{nombre_txt}.txt"
                    session.findById("wnd[1]").sendVKey(0)
                    time.sleep(0.8)
                    if session.Children.Count > 1:
                        try: session.findById("wnd[1]/tbar[0]/btn[11]").press()
                        except: pass
                    session.findById("wnd[0]/tbar[0]/okcd").text = "/n"
                    session.findById("wnd[0]").sendVKey(0)
                    st.success(f"✅ Fichero TXT ({nombre_txt}.txt) guardado correctamente.")
                except Exception as e: st.error(f"Error al generar TXT: {e}")
                finally: pythoncom.CoUninitialize()

# =====================================================================
# 🛠️ ORDER FLOW
# =====================================================================

    with tab_order_flow:
        st.subheader("🔍 Extracción de Datos (ORDER FLOW)")
        
        st.markdown("### 📋 Ficha de Identificación del Transporte (VT11 Sincronizado)")
        col_meta1, col_meta2, col_meta3, col_meta4 = st.columns(4)
        
        caja_act = f"{st.session_state.vt11_datos['tipo_caja']}-{st.session_state.vt11_datos['numero_caja']}"

            
        st.divider()
        
        btn_col1, btn_col2 = st.columns(2)
        
        with btn_col1:
            if st.button("🔍 GENERAR BITACORA FLOW", use_container_width=True, key="btn_fase_3"):
                try:
                    pythoncom.CoInitialize()
                    session = win32com.client.GetObject("SAPGUI").GetScriptingEngine.Children(0).Children(0)
                    session.findById("wnd[0]/tbar[0]/okcd").text = "/nY_LAD_65000036"
                    session.findById("wnd[0]").sendVKey(0)
                    time.sleep(1.5)
                    session.findById("wnd[0]/usr/radRB_SHINM").select()
                    session.findById("wnd[0]/usr/ctxtS_VBELN-LOW").text = str(st.session_state.entrega_num)
                    session.findById("wnd[0]/tbar[1]/btn[8]").press()
                    time.sleep(1)
                    
                    try:
                        grid = session.findById("wnd[0]/usr/cntlMY_CONTAINER/shellcont/shell")
                        grid.pressToolbarContextButton("&MB_VARIANT")
                        grid.selectContextMenuItem("&LOAD")
                        time.sleep(0.8)
                        v_grid = session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell")
                        fila_actual = int(st.session_state.sap_variante_row)
                        v_grid.currentCellRow = fila_actual
                        v_grid.selectedRows = str(fila_actual)
                        v_grid.clickCurrentCell()
                        time.sleep(1)
                    except Exception as grid_err:
                        st.error("❌ Variantes no cargadas o la fila del Layout cambió en SAP.")
                        raise grid_err
                        
                    data_list = []
                    cols_sap = ["SHIP_NUM", "CARR_DES", "CUST_NAME", "ZZAPPDATE", "ZZAPPTIME", "ZZSDCONFNUM", "PONUM", "INV_NUM", "ZZSDCONFCOMM"]
                    cols_excel = ["Transporte", "Transportista", "Cliente", "Fecha App", "Hora App", "Confirma SAP", "PO", "Factura", "Comentarios"]
                    row_count = grid.rowCount
                    if row_count > 0:
                        for row in range(row_count):
                            row_data = {}
                            for i, c in enumerate(cols_sap):
                                try: row_data[cols_excel[i]] = str(grid.getCellValue(row, c)).strip()
                                except: row_data[cols_excel[i]] = ""
                            # Adición de datos de operador, placas, ruta y No. económico para las plantil
                            data_list.append(row_data)
                        df = pd.DataFrame(data_list)
                        st.session_state.df_flow_guardado = df
                        st.session_state.datos_order_flow = df
                    session.findById("wnd[0]/tbar[0]/okcd").text = "/n"
                    session.findById("wnd[0]").sendVKey(0)
                    st.success("✅ Datos de ORDER FLOW cargados correctamente de SAP.")
                except Exception as e: st.error(f"Error Flow: {e}")
                finally: pythoncom.CoUninitialize()
        
        with btn_col2:
            if st.button("📄 GENERAR TABLA VT11", use_container_width=True, type="primary", key="btn_vt11"):
                try:
                    import pyperclip  # Librería para leer el portapapeles del sistema
                    pythoncom.CoInitialize()
                    
                    # 1. Conexión a la sesión activa de SAP
                    session = win32com.client.GetObject("SAPGUI").GetScriptingEngine.Children(0).Children(0)
                    
                    # 2. Navegación a la transacción VT11
                    session.findById("wnd[0]/tbar[0]/okcd").text = "/nVT11"
                    session.findById("wnd[0]").sendVKey(0)
                    
                    # 3. Ingreso de datos y ejecución
                    session.findById("wnd[0]/usr/ctxtK_TKNUM-LOW").text = str(st.session_state.entrega_num)
                    session.findById("wnd[0]").sendVKey(8) # F8 (Ejecutar)
                    
                    # 4. Exportación al Portapapeles (Opción [5,0])
                    session.findById("wnd[0]/mbar/menu[0]/menu[6]/menu[2]").select()
                    
                    # Esperar un instante para que aparezca la ventana emergente de selección de formato
                    time.sleep(0.5) 
                    session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[5,0]").select()
                    session.findById("wnd[1]/tbar[0]/btn[0]").press()
                    
                    # 5. Volver al menú principal para no dejar SAP trabado
                    session.findById("wnd[0]").sendVKey(3) # Atrás
                    session.findById("wnd[0]/tbar[0]/okcd").text = "/n"
                    session.findById("wnd[0]").sendVKey(0)
                    
                    # 6. Leer y procesar el texto del portapapeles
                    texto_portapapeles = pyperclip.paste()
                    
                    if not texto_portapapeles or "Shipment" not in texto_portapapeles:
                        st.error("❌ El portapapeles está vacío o no contiene el formato esperado de la VT11.")
                    else:
                        # Convertir el texto plano de SAP en filas individuales
                        lineas = [linea.strip() for linea in texto_portapapeles.split("\n")]
                        
                        datos_tabla = []
                        columnas = []
                        
                        for linea in lineas:
                            # Ignorar líneas vacías y las líneas divisorias de SAP (llenas de guiones)
                            if not linea or linea.startswith("---") or "List level" in linea or "Entries" in linea:
                                continue
                            
                            # Separar por el caracter "|" y limpiar espacios
                            elementos = [celda.strip() for celda in linea.split("|") if celda.strip() != ""]
                            
                            if not elementos:
                                continue
                                
                            # Identificar la fila de cabecera por la palabra clave 'Shipment'
                            if "Shipment" in elementos:
                                columnas = elementos
                            elif columnas: # Si ya tenemos cabecera, las siguientes filas son datos
                                # Asegurar que la fila de datos coincida en longitud con las columnas
                                if len(elementos) == len(columnas):
                                    datos_tabla.append(elementos)
                        
                        # 7. Crear DataFrame y guardar en el estado de la sesión
                        if columnas and datos_tabla:
                            df_vt11 = pd.DataFrame(datos_tabla, columns=columnas)
                            st.session_state.df_vt11_guardado = df_vt11
                            st.success("✅ Datos de la VT11 procesados y cargados correctamente.")
                        else:
                            st.warning("⚠️ No se pudieron extraer filas de datos válidas del reporte de SAP.")
                            
                except Exception as e:
                    st.error(f"🚨 Error al automatizar VT11 en SAP: {e}")
                finally:
                    pythoncom.CoUninitialize()
            
        # --- Bloque de visualización de la tabla en tu UI ---
        if st.session_state.df_vt11_guardado is not None:
            st.markdown("### 📊 Reporte Detallado de Transporte (VT11)")
            st.dataframe(st.session_state.df_vt11_guardado, use_container_width=True)
            
        st.divider()
            
        if st.session_state.df_flow_guardado is not None:
            st.dataframe(st.session_state.df_flow_guardado, use_container_width=True)
            csv = st.session_state.df_flow_guardado.to_csv(index=False).encode('utf-8')
            st.download_button("📥 Descargar CSV de Control", csv, f"Reporte_{st.session_state.entrega_num}.csv", "text/csv", use_container_width=True, key="btn_csv_control")
            st.info("💡 Dirígete a la pestaña superior '📄 FORMATOS EXCEL' para procesar e imprimir tus plantillas.")

# ==============================================================================
# 🖨️ PESTAÑA: FORMATOS EXCEL (Integración y Orquestación de Openpyxl)
# ==============================================================================
    with tab_formatos_excel:
        st.header("🖨️ Generación de Documentos")
        
        # Inicialización preventiva local de variables de control
        falta_cliente = True
        falta_destino = True
        cliente_corto = "N/A"
        destino_corto = "N/A"
        cliente_original = "N/A"
        destino_original = "N/A"
        transporte = "N/A"
        df_viaje = None
        
        # Validar si existen datos cargados previamente en la pestaña ORDER FLOW
        if st.session_state.datos_order_flow is None:
            st.warning("⚠️ No hay datos cargados en 'ORDER FLOW'. Por favor, busca primero un número de transporte en la pestaña anterior.")
        else:
            # Extraer datos de origen (crudos)
            df_viaje_completo = st.session_state.datos_order_flow
            df_viaje = df_viaje_completo[
                df_viaje_completo["Transporte"].astype(str).str.strip().ne("") & 
                df_viaje_completo["Transporte"].notna()
            ].copy()
            
            if df_viaje.empty:
                df_viaje = df_viaje_completo.copy()

            primera_fila = df_viaje.iloc[0]
            transporte = primera_fila.get("Transporte", "N/A")
            cliente_original = str(primera_fila.get("Cliente", "N/A")).strip()
            
            # Obtener Destino Original (desde df_vt11 si existe)
            df_vt11 = st.session_state.get('df_vt11_guardado', None)
            destino_original = "N/A"
            
            if df_vt11 is not None and not df_vt11.empty:
                if "Route name" in df_vt11.columns:
                    destino_original = df_vt11.iloc[0].get("Route name", "N/A")
                elif "ROUTE NAME" in df_vt11.columns.str.upper():
                    col_name = [c for c in df_vt11.columns if c.upper() == "ROUTE NAME"][0]
                    destino_original = df_vt11.iloc[0].get(col_name, "N/A")
            
            if str(destino_original).strip() == "" or str(destino_original).upper() == "N/A":
                destino_original = primera_fila.get("Destino", "N/A")
                
            destino_original = str(destino_original).strip()

            # Homologar usando las funciones seguras conectadas a session_state
            cliente_corto = obtener_cliente_corto(cliente_original)
            destino_corto = obtener_destino_corto(destino_original, cliente_original)
            
            # Validar si faltan homologar
            falta_cliente = (
                not cliente_corto or 
                cliente_corto.strip().upper() == "N/A" or 
                cliente_original.upper() == "N/A"
            )
            
            falta_destino = (
                not destino_corto or 
                destino_corto.strip().upper() == "N/A" or 
                destino_original.upper() == "N/A" or
                destino_corto.strip().upper() == destino_original.strip().upper()
            )
            
            # Interfaz de Resumen
            st.subheader(f"📋 Resumen del Transporte Seleccionado: {transporte}")
            
            col1, col2 = st.columns(2)
            with col1:
                if falta_cliente:
                    st.markdown(f"**Cliente:** :red[🔴 N/A / Sin Homologar] *(Original: {cliente_original})*")
                else:
                    st.markdown(f"**Cliente:** {cliente_corto} *(Original: {cliente_original})*")
                    
                if falta_destino:
                    st.markdown(f"**Destino:** :red[🔴 N/A / Sin Homologar] *(Original: {destino_original})*")
                else:
                    st.markdown(f"**Destino:** {destino_corto} *(Original: {destino_original})*")
                    
            with col2:
                st.markdown("**Documentos que se generarán:**")
                es_walmart = "SAM" in cliente_corto.upper() or "WALMART" in cliente_corto.upper() if (cliente_corto and not falta_cliente) else False
                es_soriana = "SORIANA" in cliente_corto.upper() or "CITY" in cliente_corto.upper() if (cliente_corto and not falta_cliente) else False
                
                if falta_cliente or falta_destino:
                    st.error("⚠️ Corrige los datos faltantes para habilitar la generación de documentos.")
                elif es_walmart:
                    st.info("📑 Bitácora de Viaje\n\n🚗 Pase Vehicular (PV)")
                elif es_soriana:
                    st.info("📑 Bitácora de Viaje\n\n🚪 Formato de Entradas y Salidas (ES)")
                else:
                    st.info("📑 Bitácora de Viaje (Cliente no requiere PV ni ES)")
                    
            st.write("---")
            
            # Formulario de homologación en caliente
            if falta_cliente or falta_destino:
                st.warning("🔍 Se han detectado datos sin homologar. Configúralos aquí mismo para continuar:")
                
                col_refrescar_izq, col_refrescar_der = st.columns([8, 2])
                with col_refrescar_der:
                    if st.button("🔄 Recargar Catálogos", use_container_width=True, key="btn_recarg_cat_local"):
                        with st.spinner("Actualizando..."):
                            recargar_catalogos_en_sesion()
                        st.toast("¡Catálogos actualizados desde el Excel!", icon="✅")
                        st.rerun()

                cols_homologar = st.columns(2)
                
                if falta_cliente:
                    with cols_homologar[0]:
                        st.markdown("### 🏢 Homologar Cliente")
                        st.text_input("Valor original en SAP:", value=cliente_original, disabled=True, key="original_cli_dis")
                        nuevo_nombre_cliente = st.text_input(
                            "Escribe el nombre corto correcto:", 
                            placeholder="Ej. SAMS, WALMART, SORIANA",
                            key="input_nuevo_cli"
                        ).strip()
                        
                        if st.button("💾 Guardar Cliente", key="btn_guardar_cli", use_container_width=True):
                            if nuevo_nombre_cliente:
                                if guardar_nuevo_registro_excel("Clientes", cliente_original, nuevo_nombre_cliente):
                                    recargar_catalogos_en_sesion()
                                    st.success("¡Cliente guardado y catálogo actualizado con éxito!")
                                    st.rerun()
                            else:
                                st.error("Por favor, escribe un nombre corto válido.")
                                
                if falta_destino:
                    with cols_homologar[1]:
                        st.markdown("### 📍 Homologar Destino")
                        st.text_input("Valor original en SAP:", value=destino_original, disabled=True, key="original_dest_dis")
                        nuevo_nombre_destino = st.text_input(
                            "Escribe el destino corto correcto:", 
                            placeholder="Ej. CEDIS GDL, CEDIS VILLAHERMOSA",
                            key="input_nuevo_dest"
                        ).strip()
                        
                        if st.button("💾 Guardar Destino", key="btn_guardar_dest", use_container_width=True):
                            if nuevo_nombre_destino:
                                if guardar_nuevo_registro_excel("Destinos", destino_original, nuevo_nombre_destino):
                                    recargar_catalogos_en_sesion()
                                    st.success("¡Destino guardado y catálogo actualizado con éxito!")
                                    st.rerun()
                            else:
                                st.error("Por favor, escribe un destino corto válido.")
                st.write("---")

            # Conversor de Excel a PDF
            def convertir_excel_a_pdf(ruta_excel, ruta_pdf):
                
                import pythoncom
                excel_app = None
                try:
                    pythoncom.CoInitialize()
                    excel_app = win32com.client.DispatchEx("Excel.Application")
                    excel_app.Visible = False
                    excel_app.DisplayAlerts = False
                    wb = excel_app.Workbooks.Open(os.path.abspath(ruta_excel))
                    for ws in wb.Worksheets:
                        ws.PageSetup.Zoom = False
                        ws.PageSetup.FitToPagesWide = 1
                        ws.PageSetup.FitToPagesTall = 1
                    wb.ExportAsFixedFormat(0, os.path.abspath(ruta_pdf))
                    wb.Close(SaveChanges=False)
                    return True
                except Exception as e:
                    st.error(f"❌ Error al convertir {os.path.basename(ruta_excel)} a PDF: {e}")
                    return False
                finally:
                    if excel_app:
                        excel_app.Quit()
                    pythoncom.CoUninitialize()

            def mostrar_boton_imprimir_pdf(ruta_pdf, label_boton="🖨️ Abrir PDF para Imprimir"):
                if os.path.exists(ruta_pdf):
                    if st.button(label_boton, key=f"btn_{ruta_pdf}", use_container_width=True):
                        try:
                            os.startfile(ruta_pdf)
                        except Exception as e:
                            st.error(f"No se pudo abrir el archivo de forma automática: {e}")

            deshabilitar_boton = falta_cliente or falta_destino
            
            # --- NUEVO: Distribución de botones en columnas ---
            col_btn_generar, col_btn_facturado = st.columns([7, 3])
            
            with col_btn_generar:
                if st.button(
                    "🚀 Generar Todos los Formatos del Viaje", 
                    use_container_width=True, 
                    disabled=deshabilitar_boton,
                    help="Configura los campos en rojo primero para poder avanzar" if deshabilitar_boton else "Comenzar proceso de generación"
                ):
                    with st.spinner("Procesando plantillas y escribiendo en archivos..."):
                        try:
                            ruta_plantillas = r"C:\Users\BP_Juancarlos Palomi\Documents\Python_SAP\python-3.13.7-embed-amd64\plantillas"
                            ruta_resultados = r"C:\Users\BP_Juancarlos Palomi\Documents\Python_SAP\python-3.13.7-embed-amd64\documentos_generados"
                            
                            if not os.path.exists(ruta_resultados):
                                os.makedirs(ruta_resultados)
                            
                            plantilla_bitacora_path = os.path.join(ruta_plantillas, "BITACORA.xlsx")
                            plantilla_es_path = os.path.join(ruta_plantillas, "ES.xlsx")
    
                            firma_gdp = st.session_state.get('firma_gdp', 'MARTIN')
                            num_transporte = str(transporte).strip() if transporte != "N/A" else "SIN_NUMERO"
                            salida_bitacora_path = os.path.join(ruta_resultados, f"BITACORA_{num_transporte}.xlsx")
                            pdf_bitacora_path = os.path.join(ruta_resultados, f"BITACORA_{num_transporte}.pdf")
    
                            documentos_impresion = []
    
                            # Generación de Bitácora
                            exito_bitacora = generar_bitacora(
                                ruta_plantilla=plantilla_bitacora_path,
                                ruta_salida=salida_bitacora_path,
                                df_vt11=df_vt11,
                                df_viaje=df_viaje,
                                firma_gdp=firma_gdp
                            )
                            
                            if exito_bitacora:
                                st.success(f"✅ ¡Bitácora generada en Excel con éxito!")
                                with st.spinner("📄 Convirtiendo Bitácora a formato PDF..."):
                                    if convertir_excel_a_pdf(salida_bitacora_path, pdf_bitacora_path):
                                        documentos_impresion.append({
                                            "ruta": pdf_bitacora_path,
                                            "nombre": f"🖨️ Imprimir Bitácora de Viaje ({num_transporte})"
                                        })
    
                            # Generación de Pase Vehicular (Walmart y Sam's)
                            if es_walmart:
                                fila_vt11_datos = df_vt11.iloc[0].to_dict() if df_vt11 is not None and not df_vt11.empty else {}
                                datos_viaje_mapeado = {
                                    "fecha_app": primera_fila.get("Fecha App", ""),
                                    "hora_app": primera_fila.get("Hora App", ""),
                                    "confirma_sap": primera_fila.get("Confirma SAP", ""),
                                    "placa_tractor": primera_fila.get("Placas", "")
                                }
    
                                lista_pos = [str(po).strip() for po in df_viaje["PO"].dropna().unique() if str(po).strip() != "" and str(po).strip().upper() != "NAN"]
                                lista_facturas = [str(fac).strip() for fac in df_viaje["Factura"].dropna().unique() if str(fac).strip() != "" and str(fac).strip().upper() != "NAN"]
    
                                generar_pase_vehicular(
                                    fila_vt11=fila_vt11_datos,
                                    datos_viaje=datos_viaje_mapeado,
                                    lista_pos=lista_pos,
                                    lista_facturas=lista_facturas,
                                    numero_viaje=num_transporte
                                )
                                
                                ruta_pases_reales = r"C:\Users\BP_Juancarlos Palomi\Documents\Python_SAP\python-3.13.7-embed-amd64\pases_generados"
                                salida_pv_path = os.path.join(ruta_pases_reales, f"PV_Viaje_{num_transporte}.xlsx")
                                pdf_pv_path = os.path.join(ruta_pases_reales, f"PV_Viaje_{num_transporte}.pdf")
                                
                                if os.path.exists(salida_pv_path):
                                    st.success(f"🚗 Pase Vehicular localizado.")
                                    with st.spinner("📄 Convirtiendo Pase Vehicular..."):
                                        if convertir_excel_a_pdf(salida_pv_path, pdf_pv_path):
                                            documentos_impresion.append({
                                                "ruta": pdf_pv_path,
                                                "nombre": f"🚗 Imprimir Pase Vehicular (Viaje: {num_transporte})"
                                            })
                                
                            # Generación de ES (Soriana y City Club)
                            elif es_soriana:
                                citas_unicas = [cita for cita in df_viaje["Confirma SAP"].dropna().unique() if str(cita).strip() != "" and str(cita).strip().upper() != "NAN"]
                                for cita in citas_unicas:
                                    filas_cita = df_viaje[df_viaje["Confirma SAP"] == cita]
                                    if filas_cita.empty:
                                        continue
                                    
                                    primera_fila_cita = filas_cita.iloc[0]
                                    cdf_destino = obtener_cdf_soriana(destino_corto)
                                    
                                    datos_es = {
                                        "transporte": num_transporte,
                                        "cita": cita,
                                        "cdf": cdf_destino,
                                        "fecha_app": primera_fila_cita.get("Fecha App", ""),
                                        "hora_app": primera_fila_cita.get("Hora App", ""),
                                        "operador": primera_fila_cita.get("Operador", ""),
                                        "placas": primera_fila_cita.get("Placas", "")
                                    }
                                    
                                    salida_es = os.path.join(ruta_resultados, f"ES_{num_transporte}_{cita}.xlsx")
                                    pdf_es = os.path.join(ruta_resultados, f"ES_{num_transporte}_{cita}.pdf")
                                    
                                    generar_es(plantilla_es_path, salida_es, datos_es)
                                    st.success(f"🚪 Formato ES generado para la cita {cita}.")
                                    
                                    with st.spinner(f"📄 Convirtiendo Formato ES a PDF..."):
                                        if convertir_excel_a_pdf(salida_es, pdf_es):
                                            documentos_impresion.append({
                                                "ruta": pdf_es,
                                                "nombre": f"🚪 Imprimir Formato ES (Cita: {cita})"
                                            })
                            
                            st.session_state["documentos_impresion_listos"] = documentos_impresion
                            st.balloons()
                            st.success(f"🎉 ¡Proceso finalizado!")
                            
                        except Exception as e:
                            st.error(f"❌ Ocurrió un error: {str(e)}")
            
            # --- NUEVO: Botón independiente "Facturado" ---
            # --- NUEVO: Botón independiente "Facturado" (CORREGIDO PARA DESBLOQUEAR) ---
            with col_btn_facturado:
                num_transporte = str(transporte).strip() if transporte != "N/A" else "SIN_NUMERO"
                
                # Modificado: Ahora solo se bloquea si de verdad no hay ningún número de viaje seleccionado
                deshabilitar_facturado = num_transporte == "SIN_NUMERO" or num_transporte == ""
                
                if st.button(
                    "📈 Marcar como Facturado", 
                    use_container_width=True, 
                    type="secondary",
                    disabled=deshabilitar_facturado,
                    help="Cambiar el estatus del viaje actual a FACTURADO en Google Sheets sin generar formatos."
                ):
                    with st.spinner("🔄 Actualizando estatus en Google Sheets..."):
                        if actualizar_status_en_sheets(num_transporte):
                            st.toast(f"¡Viaje {num_transporte} actualizado a FACTURADO! ✅", icon="📈")
                            st.success(f"📈 Se cambió el estatus del viaje {num_transporte} a FACTURADO.")

            # Panel de Impresión Rápida
            if "documentos_impresion_listos" in st.session_state and st.session_state["documentos_impresion_listos"]:
                st.write("---")
                st.subheader("🖨️ Panel de Impresión Rápida")
                
                documentos = st.session_state["documentos_impresion_listos"]
                cols_print = st.columns(len(documentos))
                for idx, doc in enumerate(documentos):
                    with cols_print[idx]:
                        mostrar_boton_imprimir_pdf(doc["ruta"], doc["nombre"])

# ==============================================================================
# 🖨️ PESTAÑA: SPOOL
# ==============================================================================

    with tab_spool:
        st.header("🖨️ Extractor de Spool")
        temp_path = os.path.join(os.path.expanduser("~"), "temp_spool.txt")
        if st.button("🚀 Extraer Fichero Local", type="primary", key="spool_local"):
            pythoncom.CoInitialize()
            try:
                session = win32com.client.GetObject("SAPGUI").GetScriptingEngine.Children(0).Children(0)
                session.findById("wnd[0]/tbar[0]/okcd").text = "/nSP02"
                session.findById("wnd[0]").sendVKey(0)
                time.sleep(1)
                session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").setCurrentCell(0, "RQIDENT")
                session.findById("wnd[0]/mbar/menu[1]/menu[0]/menu[5]/menu[2]").select()
                session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[0,0]").select()
                session.findById("wnd[1]/tbar[0]/btn[0]").press()
                session.findById("wnd[1]/usr/ctxtDY_PATH").text = os.path.dirname(temp_path)
                session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = os.path.basename(temp_path)
                session.findById("wnd[1]/tbar[0]/btn[0]").press()
                time.sleep(2)
                if os.path.exists(temp_path):
                    with open(temp_path, 'r', encoding='latin-1') as f: contenido = f.readlines()
                    st.dataframe(pd.DataFrame(contenido, columns=["Contenido Spool"]), use_container_width=True)
                    os.remove(temp_path)
            except Exception as e: st.error(f"❌ Error: {e}")
            finally: pythoncom.CoUninitialize()
                
    st.divider()
    st.caption("Nota: Requiere mantener tu ventana activa de SAP GUI abierta de fondo.")


# =====================================================================
# ⚡ HERRAMIENTA 2: CONTROL TRÁNSITOS MX21 (Código Completo)
# =====================================================================
def fase_1_traer_vista_previa(shipments_texto):
    pythoncom.CoInitialize()
    try:
        lista_shipments = [s.strip() for s in shipments_texto.split("\n") if s.strip()]
        if not lista_shipments:
            return False, "Error: No ingresaste ningún número de Shipment.", None
        texto_para_copiar = "\r\n".join(lista_shipments)
        subprocess.run("clip", input=texto_para_copiar, text=True, check=True)
        sap_gui = win32com.client.GetObject("SAPGUI")
        if not sap_gui: return False, "No se pudo obtener el objeto SAPGUI. ¿Está abierto SAP?", None
        session = sap_gui.GetScriptingEngine.Children(0).Children(0)
        st.toast("Conectado a SAP. Extrayendo vista previa...", icon="📊")
        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").text = "/nY_LAD_65000127"
        session.findById("wnd[0]").sendVKey(0)
        time.sleep(1.5) 
        session.findById("wnd[0]/usr/rad%ALV").setFocus()
        session.findById("wnd[0]/usr/rad%ALV").select()
        session.findById("wnd[0]/usr/btn%_SP$00002_%_APP_%-VALU_PUSH").press()
        time.sleep(0.5)
        session.findById("wnd[1]/tbar[0]/btn[24]").press()
        session.findById("wnd[1]").sendVKey(8)  
        session.findById("wnd[0]").sendVKey(8)  
        time.sleep(1.5) 
        session.findById("wnd[0]/tbar[1]/btn[33]").press()
        shell_layout = session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell")
        shell_layout.currentCellRow = 175
        shell_layout.firstVisibleRow = 163
        shell_layout.selectedRows = "175"
        shell_layout.clickCurrentCell()
        time.sleep(1.5) 
        session.findById("wnd[0]/tbar[1]/btn[45]").press()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[4,0]").select()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[4,0]").setFocus()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        time.sleep(1.2) 
        df_sap = pd.read_clipboard(sep='\t', dtype=str).fillna("")
        if not df_sap.empty:
            df_sap = df_sap.map(lambda x: str(x).replace('|', '').strip() if isinstance(x, str) else x)
            df_sap = df_sap[~df_sap.iloc[:, 0].str.contains(r'\*|---', na=False)] 
            return True, "¡Datos de Shipments capturados! Revisa la tabla abajo antes de confirmar.", df_sap
        else: return False, "SAP terminó el proceso pero no se detectaron datos legibles.", None
    except Exception as e: return False, f"Error en Vista Previa Fase 1: {str(e)}", None
    finally: pythoncom.CoUninitialize()

def fase_1_confirmar_y_finalizar():
    pythoncom.CoInitialize()
    try:
        sap_gui = win32com.client.GetObject("SAPGUI")
        session = sap_gui.GetScriptingEngine.Children(0).Children(0)
        session.findById("wnd[0]/tbar[0]/btn[86]").press()
        session.findById("wnd[1]/tbar[0]/btn[13]").press()
        time.sleep(1) 
        session.findById("wnd[0]").sendVKey(3)
        session.findById("wnd[0]").sendVKey(3)
        return True, "¡Procesado e impreso correctamente en SAP!"
    except Exception as e: return False, f"Error al finalizar en SAP: {str(e)}"
    finally: pythoncom.CoUninitialize()

def ejecutar_fase_2(shipment_id, p_date, p_time_raw, d_date, d_time_raw):
    pythoncom.CoInitialize()
    try:
        p_time_raw = "".join(filter(str.isdigit, p_time_raw))
        d_time_raw = "".join(filter(str.isdigit, d_time_raw))
        p_time = f"{p_time_raw[:2]}:{p_time_raw[2:]}" if len(p_time_raw) == 4 else p_time_raw
        d_time = f"{d_time_raw[:2]}:{d_time_raw[2:]}" if len(d_time_raw) == 4 else d_time_raw
        sap_gui = win32com.client.GetObject("SAPGUI")
        if not sap_gui: return "No se pudo obtener el objeto SAPGUI."
        session = sap_gui.GetScriptingEngine.Children(0).Children(0)
        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").text = "/nVT02N"
        session.findById("wnd[0]").sendVKey(0)
        time.sleep(1)
        session.findById("wnd[0]/usr/ctxtVTTK-TKNUM").text = str(shipment_id)
        session.findById("wnd[0]").sendVKey(0)
        time.sleep(1)
        session.findById("wnd[0]/usr/tabsHEADER_TABSTRIP2/tabpTABS_OV_TE").select()
        sub_path = "wnd[0]/usr/tabsHEADER_TABSTRIP2/tabpTABS_OV_TE/ssubG_HEADER_SUBSCREEN2:SAPMV56A:1035/"
        session.findById(sub_path + "ctxtVTTK-TNDR_ERPD").text = str(p_date)
        session.findById(sub_path + "ctxtVTTK-TNDR_ERPT").text = str(p_time)
        session.findById(sub_path + "ctxtVTTK-TNDR_LTPD").text = str(p_date)
        session.findById(sub_path + "ctxtVTTK-TNDR_LTPT").text = str(p_time)
        session.findById(sub_path + "ctxtVTTK-TNDR_ERDD").text = str(d_date)
        session.findById(sub_path + "ctxtVTTK-TNDR_ERDT").text = str(d_time)
        session.findById(sub_path + "ctxtVTTK-TNDR_LTDD").text = str(d_date)
        session.findById(sub_path + "ctxtVTTK-TNDR_LTDT").text = str(d_time)
        session.findById(sub_path + "ctxtVTTK-TNDR_LTDT").setFocus()
        for _ in range(5): session.findById("wnd[0]").sendVKey(0)
        session.findById("wnd[0]").sendVKey(11)
        time.sleep(1)
        try: session.findById("wnd[1]").sendVKey(0)
        except: pass 
        return f"¡Shipment {shipment_id} actualizado (Horas: {p_time} / {d_time})!"
    except Exception as e: return f"Error en Fase 2: {str(e)}"
    finally: pythoncom.CoUninitialize()

def ejecutar_fase_3():
    pythoncom.CoInitialize()
    try:
        ruta_escritorio = r"C:\Users\BP_Israel H Higareda\Desktop\planeacion 2026"
        archivo_final = os.path.join(ruta_escritorio, "TRANSITO.txt")
        if os.path.exists(archivo_final):
            try:
                os.remove(archivo_final)
                st.toast("Archivo TRANSITO.txt viejo eliminado con éxito.", icon="🗑️")
            except Exception as file_err: return False, f"No se pudo eliminar el archivo viejo: {str(file_err)}", ""
        sap_gui = win32com.client.GetObject("SAPGUI")
        if not sap_gui: return False, "SAPGUI no encontrado. Abre SAP primero.", ""
        session = sap_gui.GetScriptingEngine.Children(0).Children(0)
        st.toast("Abriendo reporte Y_LAD_65000127...", icon="🚀")
        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").text = "/nY_LAD_65000127"
        session.findById("wnd[0]").sendVKey(0)
        time.sleep(1.2)
        session.findById("wnd[0]/usr/ctxtSP$00001-LOW").text = "mx*"
        session.findById("wnd[0]/usr/ctxtSP$00004-LOW").text = "mx21"
        session.findById("wnd[0]/usr/txtSP$00009-LOW").setFocus()
        session.findById("wnd[0]").sendVKey(2) 
        time.sleep(0.4)
        session.findById("wnd[1]/usr/cntlOPTION_CONTAINER/shellcont/shell").currentCellColumn = "TEXT"
        session.findById("wnd[1]/usr/cntlOPTION_CONTAINER/shellcont/shell").doubleClickCurrentCell()
        session.findById("wnd[0]/usr/txtSP$00010-LOW").setFocus()
        session.findById("wnd[0]").sendVKey(2)
        time.sleep(0.4)
        session.findById("wnd[1]/usr/cntlOPTION_CONTAINER/shellcont/shell").setCurrentCell(1, "TEXT")
        session.findById("wnd[1]/usr/cntlOPTION_CONTAINER/shellcont/shell").selectedRows = "1"
        session.findById("wnd[1]/usr/cntlOPTION_CONTAINER/shellcont/shell").doubleClickCurrentCell()
        session.findById("wnd[0]/usr/txtSP$00010-LOW").text = "1.000"
        session.findById("wnd[0]/usr/ctxtSP$00012-LOW").text = "6"
        session.findById("wnd[0]/usr/ctxtSP$00012-HIGH").text = "7"
        hoy = datetime.now()
        hace_un_mes = hoy - timedelta(days=30)
        f_low_sap = hace_un_mes.strftime("%d%m%Y")
        f_high_sap = hoy.strftime("%d%m%Y")
        session.findById("wnd[0]/usr/ctxtSP$00013-LOW").text = f_low_sap
        session.findById("wnd[0]/usr/ctxtSP$00013-HIGH").text = f_high_sap
        session.findById("wnd[0]/usr/rad%ALV").setFocus()
        session.findById("wnd[0]/usr/rad%ALV").select()
        session.findById("wnd[0]").sendVKey(8)
        time.sleep(2)
        session.findById("wnd[0]/tbar[1]/btn[33]").press()
        shell_layout = session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell")
        shell_layout.firstVisibleRow = 275
        shell_layout.currentCellRow = 175
        shell_layout.firstVisibleRow = 173
        shell_layout.selectedRows = "175"
        shell_layout.clickCurrentCell()
        time.sleep(1.5)
        session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").currentCellRow = -1
        session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").selectColumn("VTTK-TKNUM")
        session.findById("wnd[0]/tbar[1]/btn[45]").press()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").select()
        session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").setFocus()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        session.findById("wnd[1]/usr/ctxtDY_PATH").text = ruta_escritorio
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "TRANSITO.txt"
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        time.sleep(1.5)
        session.findById("wnd[0]").sendVKey(3)
        session.findById("wnd[0]").sendVKey(3)
        if os.path.exists(archivo_final):
            with open(archivo_final, "r", encoding="utf-8", errors="ignore") as f: contenido = f.read()
            shipments_limpios = []
            lineas = contenido.split("\n")
            for line in lineas:
                match = re.search(r'\b([68]\d{9})\b', line)
                if match: shipments_limpios.append(match.group(1))
            return True, "¡Paso 1 Completado! Shipments extraídos con éxito.", "\n".join(shipments_limpios)
        else: return False, "Error: No se detectó la creación física de TRANSITO.txt.", ""
    except Exception as e: return False, f"Error Crítico en la extracción: {str(e)}", ""
    finally: pythoncom.CoUninitialize()

def ejecutar_fase_3_personalizada(shipments_pantalla):
    pythoncom.CoInitialize()
    try:
        lista_shipments = [s.strip() for s in shipments_pantalla.split("\n") if s.strip()]
        if not lista_shipments: return False, "Error: No hay Shipments en el cuadro de la derecha para procesar."
        texto_para_sap = "\r\n".join(lista_shipments)
        subprocess.run("clip", input=texto_para_sap, text=True, check=True)
        ruta_escritorio = r"C:\Users\BP_Israel H Higareda\Desktop\planeacion 2026"
        archivo_csv = os.path.join(ruta_escritorio, "TRANSITOS.csv")
        if os.path.exists(archivo_csv):
            try:
                os.remove(archivo_csv)
                st.toast("Archivo TRANSITOS.csv previo eliminado con éxito.", icon="🗑️")
            except Exception as fe: return False, f"No se pudo limpiar el archivo csv viejo: {str(fe)}"
        sap_gui = win32com.client.GetObject("SAPGUI")
        if not sap_gui: return False, "SAPGUI no encontrado. Asegúrate de tener SAP abierto."
        session = sap_gui.GetScriptingEngine.Children(0).Children(0)
        st.toast("Iniciando tu macro personalizada en SAP...", icon="🤖")
        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/usr/cntlIMAGE_CONTAINER/shellcont/shell/shellcont[0]/shell").doubleClickNode("F00002")
        time.sleep(1)
        session.findById("wnd[0]/usr/btn%_K_TKNUM_%_APP_%-VALU_PUSH").press()
        time.sleep(0.5)
        session.findById("wnd[1]/tbar[0]/btn[24]").press() 
        time.sleep(0.5)
        session.findById("wnd[1]").sendVKey(8)
        session.findById("wnd[0]/tbar[1]/btn[8]").press() # Ejecutar reporte F8
        time.sleep(2)
        session.findById("wnd[0]/mbar/menu[3]/menu[0]/menu[1]").select()
        time.sleep(0.5)
        session.findById("wnd[1]/usr").verticalScrollbar.position = 641
        session.findById("wnd[1]/usr").verticalScrollbar.position = 640
        session.findById("wnd[1]/usr").verticalScrollbar.position = 639
        session.findById("wnd[1]/usr").verticalScrollbar.position = 638
        session.findById("wnd[1]/usr").verticalScrollbar.position = 637
        session.findById("wnd[1]/tbar[0]/btn[71]").press()
        time.sleep(0.5)
        session.findById("wnd[2]/usr/txtRSYSF-STRING").text = "/ZZBCXC"
        session.findById("wnd[2]/usr/txtRSYSF-STRING").caretPosition = 7
        session.findById("wnd[2]/tbar[0]/btn[0]").press()
        time.sleep(0.5)
        session.findById("wnd[3]/usr/lbl[1,2]").setFocus()
        session.findById("wnd[3]/usr/lbl[1,2]").caretPosition = 5
        session.findById("wnd[3]").sendVKey(2)
        time.sleep(0.5)
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        time.sleep(1)
        session.findById("wnd[0]/mbar/menu[0]/menu[7]/menu[0]").select()
        time.sleep(0.5)
        session.findById("wnd[1]/usr/ssubSUB_CONFIGURATION:SAPLSALV_GUI_CUL_EXPORT_AS:0512/txtGS_EXPORT-FILE_NAME").text = "TRANSITOS.CSV"
        session.findById("wnd[1]/usr/ssubSUB_CONFIGURATION:SAPLSALV_GUI_CUL_EXPORT_AS:0512/txtGS_EXPORT-FILE_NAME").caretPosition = 14
        session.findById("wnd[1]/tbar[0]/btn[20]").press()
        time.sleep(0.5)
        session.findById("wnd[1]/usr/ctxtDY_PATH").setFocus()
        session.findById("wnd[1]/usr/ctxtDY_PATH").caretPosition = 0
        session.findById("wnd[1]").sendVKey(4)
        time.sleep(0.5)
        session.findById("wnd[2]/usr/ctxtDY_PATH").text = ruta_escritorio
        session.findById("wnd[2]/usr/ctxtDY_FILENAME").text = "TRANSITOS.csv"
        session.findById("wnd[2]/usr/ctxtDY_FILENAME").caretPosition = 10
        session.findById("wnd[2]/tbar[0]/btn[0]").press()
        time.sleep(0.5)
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        time.sleep(1.5)
        session.findById("wnd[0]").sendVKey(3)
        session.findById("wnd[0]").sendVKey(3)
        if os.path.exists(archivo_csv): return True, "¡Paso 2 Completado! Se generó el archivo estructurado."
        else: return False, "La macro terminó pero el archivo TRANSITOS.csv no se localiza en el escritorio."
    except Exception as e: return False, f"Error en la macro de comandos SAP: {str(e)}"
    finally: pythoncom.CoUninitialize()

def render_control_transitos():
    st.title("⚡ Sistema Central de Transitos MX21 (SAP)")
    tab1, tab2, tab3 = st.tabs(["📊 Fase 1: GR", "🚚 Fase 2: TENDER", "🔍 Fase 3: TRANSITOS"])

    if "fase1_df" not in st.session_state: st.session_state.fase1_df = None
    if "fase3_lista" not in st.session_state: st.session_state.fase3_lista = ""
    if "fase3_listo_para_mostrar" not in st.session_state: st.session_state.fase3_listo_para_mostrar = False

    with tab1:
        st.subheader("Extracción e Impresión Segura ALV")
        txt_shipments = st.text_area("📋 Pega tus Shipments (Uno por línea):", height=150, key="txt1")
        if st.button("🔍 1. Traer Vista Previa desde SAP", key="b_f1_1", use_container_width=True):
            if not txt_shipments.strip(): st.warning("Por favor ingresa Shipments.")
            else:
                with st.spinner("Conectando con SAP..."):
                    ok, msg, df = fase_1_traer_vista_previa(txt_shipments)
                    if ok: st.session_state.fase1_df = df; st.info(msg)
                    else: st.error(msg); st.session_state.fase1_df = None

        if st.session_state.fase1_df is not None:
            st.dataframe(st.session_state.fase1_df, use_container_width=True)
            if st.button("✅ 2. SÍ, Datos Correctos. DAR ACEPTAR EN SAP", type="primary", key="b_f1_2", use_container_width=True):
                with st.spinner("Aprobando en SAP..."):
                    ok_f, msg_f = fase_1_confirmar_y_finalizar()
                    if ok_f: st.success(msg_f); st.session_state.fase1_df = None
                    else: st.error(msg_f)

    with tab2:
        st.subheader("TENDER")
        input_shipment = st.text_input("📦 Número de Shipment:", placeholder="Ej: 6500524052", max_chars=10)
        st.divider()
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**📅 Planificación de Recolección (Pickup)**")
            pickup_date = st.text_input("Fecha de Carga:", value="27.05.2026")
            pickup_time = st.text_input("Hora de Carga (Solo números):", value="1135")
        with c2:
            st.markdown("**📅 Planificación de Entrega (Delivery)**")
            delivery_date = st.text_input("Fecha de Destino:", value="27.05.2026")
            delivery_time = st.text_input("Hora de Destino (Solo números):", value="1627")
        st.divider()
        if st.button("💾 Actualizar y Guardar en SAP", key="b_f2", use_container_width=True):
            if not input_shipment: st.warning("Falta número de Shipment.")
            else:
                with st.spinner("Guardando en SAP..."):
                    st.success(ejecutar_fase_2(input_shipment, pickup_date, pickup_time, delivery_date, delivery_time))

    with tab3:
        st.subheader("Fase 3: Flujo Completo de Automatización Logística")
        col_izq, col_der = st.columns([1.2, 1])
        with col_izq:
            st.markdown("INTRACOMPANY Y VT11")
            st.caption("Paso 1: Generar la lista base desde el reporte general de SAP.")
            if st.button("📋 Botón 1: Descargar Tránsito y Extraer Shipments", key="btn_run_f3_orig", use_container_width=True):
                st.session_state.fase3_listo_para_mostrar = False 
                with st.spinner("Ejecutando Intracompany..."):
                    ok_f3, msg_f3, lista_f3 = ejecutar_fase_3()
                    if ok_f3: st.session_state.fase3_lista = lista_f3; st.success(msg_f3)
                    else: st.error(msg_f3)
            st.divider()
            st.caption("Paso 2: Usar los Shipments de la derecha para correr tu macro específica de Layout.")
            if st.button("🚀 Botón 2: Correr y Guardar TRANSITOS.csv", key="btn_run_f3_custom", use_container_width=True):
                if not st.session_state.fase3_lista.strip(): st.error("Error: No hay Shipments en el cuadro de la derecha para procesar.")
                else:
                    with st.spinner("Inyectando transportes y aplicando Layout /ZZBCXC..."):
                        ok_c, msg_c = ejecutar_fase_3_personalizada(st.session_state.fase3_lista)
                        if ok_c: st.session_state.fase3_listo_para_mostrar = True; st.success(msg_c)
                        else: st.session_state.fase3_listo_para_mostrar = False; st.error(msg_c)
                        
        with col_der:
            st.markdown("### 📋 Shipments en Pantalla")
            st.caption("Los números aquí cargados se usarán automáticamente para alimentar el Botón 2:")
            st.text_area("Lista de trabajo activa:", value=st.session_state.fase3_lista, height=250, key="fase3_lista_area")


# =====================================================================
# 🎛️ ORQUESTADOR DE RENDERIZADO PRINCIPAL
# =====================================================================
if opcion == "🚛 SAP Logistics Master":
    render_logistica_master()
elif opcion == "⚡ Control Tránsitos MX21":
    render_control_transitos()